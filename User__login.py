import openpyxl as xl
import re
from datetime import datetime
import datetime as xyz
import pywhatkit
from pyfiglet import Figlet
from termcolor import colored
from PIL import Image

wb = xl.load_workbook('transactions.xlsx')
sheet1 = wb['Sheet1']
sheet2 = wb['Sheet2']
line = colored("-", 'magenta')


# Opening Message.

def opening_message():
    first_line = Figlet(font="big")
    first_line = colored(first_line.renderText('Welcome to Manchester Royal.'), 'magenta')
    img = Image.open('Manchester Royal Front Elavation.png')
    img.show()
    opening_msg = colored(
        "\nStanding proudly at the center of Gulberg Greens, this development is poised to Gulbergs most exciting "
        "living and working destination. \n", 'magenta')

    print(first_line.center(len(opening_msg), "-") + "\n" + opening_msg.center(len(opening_msg)))


def goodbye_msg():
    goodbye_msg_display = Figlet(font='big')
    goodbye_msg_display = goodbye_msg_display.renderText("Thank    you    for    visiting  .    Allah    Hafiz")
    print(colored(goodbye_msg_display, 'magenta'))


def xl_defaulter_or_ok():
    current_date = datetime.now().date()  # Gets today's Date
    for row in range(2, sheet2.max_row + 1):
        paid_percentage_cell = sheet2.cell(row, 10)
        payment_made_status_cell = sheet2.cell(row, 13)
        to_be_made_payment_date_cell = sheet2.cell(row, 12)
        if paid_percentage_cell.value != "0" and paid_percentage_cell.value != "100":
            to_be_made_payment_date_cell.value = to_be_made_payment_date_cell.value.date()
            if current_date > to_be_made_payment_date_cell.value:

                payment_made_status_cell.value = "defaulter"

            else:
                payment_made_status_cell.value = "OK"

    wb.save("transactions.xlsx")


def loading_xl_file():
    return wb, sheet1, sheet2


def signup():
    print(colored("Welcome to Signup\n", "green").center(134))
    row_signup = sheet1.max_row + 1
    # Username

    username_check = []
    for x in range(2, sheet1.max_row + 1):
        username_value = sheet1.cell(x, 1)
        username_check.append(username_value.value)
    # Turns list in lowercase

    for i in range(len(username_check)):
        username_check[i] = username_check[i].lower()

    # Username along with uniqueness

    xl_username = sheet1.cell(row_signup, 1)
    xl_username.value = input(colored("Enter Username you would like to keep.\n", "yellow")).lower()
    while xl_username.value in username_check:
        xl_username = sheet1.cell(row_signup, 1)
        xl_username.value = input(colored("Username already Exists Try another username.\n", "red")).lower()

    print(colored("Username Set Successfully.\n", "green").center(134))

    # Password Input
    xl_password = sheet1.cell(row_signup, 2)
    xl_password.value = input(colored("Set an 8 or more character alphanumeric passcode\n", 'yellow'))

    # Password Length along with Alphanumeric check
    while len(xl_password.value) < 8 or not xl_password.value.isalnum():
        xl_password = sheet1.cell(row_signup, 2)
        xl_password.value = input(colored("Password is invalid (Length less than 8 or special Charachter used). "
                                          "Re-enter\n", 'red'))

    print(colored("Password Set Successfully.", "green").center(134))

    # Type is Buyer As Admin waghaira preset
    xl_account_type = sheet1.cell(row_signup, 3)
    xl_account_type.value = "buyer"

    # Name Entered
    xl_name = sheet1.cell(row_signup, 4)
    xl_name.value = input(colored("Enter your name.\n", 'yellow')).lower()

    name_checker_joined = xl_name.value.replace(" ", "")

    # Name Check
    while not name_checker_joined.isalpha():
        xl_name = sheet1.cell(row_signup, 4)
        xl_name.value = input(colored("Only use letters in your name. Re-enter.\n", 'red')).lower()
        name_checker_joined = xl_name.value.replace(" ", "")

    # Cnic length
    xl_cnic = sheet1.cell(row_signup, 5)
    xl_cnic.value = input(colored("Enter your CNIC.\n", 'yellow'))
    while not re.search('^[0-9]{5}-[0-9]{7}-[0-9]$', xl_cnic.value):
        xl_cnic = sheet1.cell(row_signup, 5)
        xl_cnic.value = input(colored("Invalid Format. Re-enter\n", 'red'))

    print(colored("CNIC Accepted Successfully.", 'green').center(134))

    # Phone number
    xl_phone_number = sheet1.cell(row_signup, 6)

    xl_phone_number.value = input(colored("Enter Your Phone number in format [+92000-0000000].\n", 'yellow'))
    while not re.search('^[+]92[0-9]{3}-[0-9]{7}$', xl_phone_number.value):
        xl_phone_number = sheet1.cell(row_signup, 5)
        xl_phone_number.value = input(colored("Invalid Format. Re-enter\n", 'red'))

    print(colored(
        f"You Have Successfully Signed up", 'green').title().center(134) + "\n" + colored(f"Welcome {xl_name.value}",
                                                                                          'cyan').title().center(134))
    # saving xl file
    wb.save('transactions.xlsx')

    return xl_username.value, xl_password.value, xl_account_type.value, xl_name.value, xl_cnic.value, xl_phone_number.value


# User name exists or not
def username(xl_username, xl_password):
    username_check = []
    password_check = []

    for x in range(2, sheet1.max_row + 1):
        username_value = sheet1.cell(x, 1)
        username_check.append(username_value.value)
        password_value = sheet1.cell(x, 2)
        password_check.append(password_value.value)

    # Turns list in lowercase
    for i in range(len(username_check)):
        username_check[i] = username_check[i].lower()

    # If doesnt exist
    while xl_username.lower() not in username_check:
        xl_username = input(colored("This account Does not exist. Try again\n", 'red')).lower()

    index_username = username_check.index(xl_username.lower())

    while not password_check[index_username] == xl_password:
        xl_password = input(colored("Wrong Passcode. Re-try\n", 'red'))

    print(colored("Password Accepted", 'green').center(134))
    xl_username = sheet1.cell(index_username + 2, 1)
    account_type = sheet1.cell(index_username + 2, 3)
    name_displayed = sheet1.cell(index_username + 2, 4)
    cnic_user = sheet1.cell(index_username + 2, 5)
    phone_no_user = sheet1.cell(index_username + 2, 6)

    wb.save('transactions.xlsx')

    print(colored(f"You Have Successfully Logged in", 'green').title().center(
        134) + "\n" + colored(f"Welcome {name_displayed.value}", 'cyan').title().center(134))
    return xl_username.value, xl_password, account_type.value, name_displayed.value, cnic_user.value, phone_no_user.value


# Buyer Menu List
def buyer_choice():
    print(line * 136)
    print(colored("(1) Showing of Available Units along with buying.".title(), "cyan").center(136) + "\n" +
          colored("(2) View Status of Bought Units".title(), "cyan").center(136) + "\n" +
          colored("(3) Make a Payment".title(), "cyan").center(136) + "\n" +
          colored("(4) Questions for Human Resource".title(), 'cyan').center(136) + "\n" +
          colored("(5) Quit".title(), 'cyan').center(136))
    print(line * 136)
    choice = input(" " * 67)

    while not re.search('^[1-5]$', choice):
        choice = input(colored("Invalid Choice re-enter: ", 'red'))
    return choice


def finance_choice():
    print(line * 136)

    print(colored("(1) Showing Payment status of defaulters.".title(), 'cyan').center(136) + "\n" +
          colored("(2) Update Monthly Payment Receipt".title(), 'cyan').center(136) + "\n" +
          colored("(3) Send mgs to a Defaulter".title(), 'cyan').center(136) + "\n" +
          colored("(4) Quit".title(), 'cyan').center(136))

    print(line * 136)

    choice = input(" " * 67)

    while not re.search('^[1-4]$', choice):
        choice = input(colored("Invalid Choice re-enter: ", 'red'))
    return choice


def hr_choice():
    print(line * 136)

    print(colored("(1) Any Question received?".title(), 'cyan').center(136) + "\n" + colored(
        "(2) Send Plaza Update".title(), 'cyan').center(136)), "\n" + colored("(3) Quit".title(), 'cyan').center(136)

    print(line * 136)

    choice = input(" " * 67)

    while not re.search('^[1-3]$', choice):
        choice = input(colored("Invalid Choice re-enter: ", 'red'))
    return choice


def admin_choice():
    print(line * 136)

    print(colored("(1) View Finance Report.".title(), 'cyan').center(136) + "\n" +
          colored("(2) Grant a Refund".title(), 'cyan').center(136) + "\n" +
          colored("(3) Look at Inventory".title(), 'cyan').center(136) + "\n" +
          colored("(4) Quit".title(), 'cyan').center(136))
    print(line * 136)

    choice = input(" " * 67)

    while not re.search('^[1-4]$', choice):
        choice = input(colored("Invalid Choice re-enter: ", 'red'))
    return choice


def refund_granted():
    sold_product_list = []
    for row in range(2, sheet2.max_row + 1):
        sold_product = sheet2.cell(row, 7)

        sold_product_list.append(sold_product.value)

    print()
    for conversion in range(len(sold_product_list)):
        if sold_product_list[conversion] == "avalible":
            sold_product_list[conversion] = 0
        else:
            sold_product_list[conversion] = 1
    i = 0
    about = ['Unit Type:', 'Unit Floor:', 'Unit Size:', 'Quarterly Installment:', 'Total Price:', 'Unit is']
    index = 0
    x = 1
    s_no_value_lst = []
    potential_sales_single_element = []

    for rows in range(2, sheet2.max_row + 1):
        if sold_product_list[index] == 1:
            print(colored(f"Option {x}", 'magenta'))
            x += 1

        for column in range(2, 8):

            if sold_product_list[index] == 1:
                s_no_cell = sheet2.cell(rows, 1)
                s_no_value = s_no_cell.value
                if s_no_value not in s_no_value_lst:
                    s_no_value_lst.append(s_no_value)
                available_options_cell = sheet2.cell(rows, column)
                available_options_value = available_options_cell.value
                print(f"{about[i]} {available_options_value}".title())
                potential_sales_single_element.append(available_options_value)
                i += 1
        if sold_product_list[index] == 1:
            print()

        index += 1

        i = 0

    print(colored("Select Option No To Be Refunded".title(), 'yellow').center(136, "-"))
    option_no = int(input(" " * 67))

    while option_no <= 0 or option_no > len(s_no_value_lst):
        print(colored("No Option Like That Exists. Re-enter".title(), 'red').center(136, "-"))
        option_no = int(input(" " * 67))
    xl_row_num = int(s_no_value_lst[option_no - 1]) + 1

    status_cell = sheet2.cell(xl_row_num, 7)
    status_cell.value = "avalible"
    sold_to_cell = sheet2.cell(xl_row_num, 8)
    sold_to_cell.value = "avalible"
    paid_amount_cell = sheet2.cell(xl_row_num, 9)
    dis_refund = paid_amount_cell.value
    paid_amount_cell.value = 0
    sold_to_percentage = sheet2.cell(xl_row_num, 10)
    sold_to_percentage.value = "0"
    payment_made = sheet2.cell(xl_row_num, 11)
    payment_made.value = ""
    payment_to_be_made = sheet2.cell(xl_row_num, 12)
    payment_to_be_made.value = ""
    payment_status = sheet2.cell(xl_row_num, 13)
    payment_status.value = ""
    wb.save("transactions.xlsx")
    print(colored(f"Payment of Rs.{dis_refund} has been Refunded.".title(), 'green').center(136, "-"))


def question_checker():
    txt_file = open("hr reply.txt", "r+")
    text_string_content = txt_file.read()
    if text_string_content == "":
        print(colored("No Msgs to check.", 'green'))
    else:
        txt_lst = text_string_content.split(",")
        txt_lst.remove("")
        txt_lst = list(set(txt_lst))
        print(colored("You have to reply to", 'red').center(136))
        for i in range(len(txt_lst)):
            print(colored(f"{i + 1}.{txt_lst[i]}", 'magenta').center(136))
        print(colored("Have you replied to them all?. Y or N?", "yellow").center(136))
        check_replied = input(" " * 67).upper()

        while check_replied != "N" and check_replied != "Y":
            print(colored("Invalid. Enter Y or N.", 'red').center(136))
            check_replied = input(" " * 67).upper()

        if check_replied == "Y":
            txt_file.truncate(0)
    txt_file.close()


def defaulter_display():
    x = 1
    defaulter_name_lst = []
    defaulter_username_lst = []
    defaulter_phone_no_lst = []

    for rows in range(2, sheet2.max_row + 1):

        for columns in range(2, 8):
            payment_made_cell = sheet2.cell(rows, 13)
            payment_made_cell_value = payment_made_cell.value
        if payment_made_cell_value == "defaulter":
            defaulter_username_cell = sheet2.cell(rows, 8)
            defaulter_username_lst.append(defaulter_username_cell.value)
            print(colored(f"Defaulter {x}", "red"))
            print(colored(defaulter_username_cell.value, 'magenta'))
            x += 1
            for row in range(5, sheet1.max_row + 1):

                username_cell = sheet1.cell(row, 1)
                username_cell_value = username_cell.value
                if username_cell_value == defaulter_username_cell.value:
                    defaulter_name_cell = sheet1.cell(row, 4)
                    defaulter_name_lst.append(defaulter_name_cell.value)

                    defaulter_phone_no_cell = sheet1.cell(row, 6)
                    defaulter_phone_no_cell_value = defaulter_phone_no_cell.value
                    defaulter_phone_no_cell_value = defaulter_phone_no_cell_value.split("-")
                    defaulter_phone_no_cell_value = "".join(defaulter_phone_no_cell_value)
                    defaulter_phone_no_lst.append(defaulter_phone_no_cell_value)

                    print(colored(defaulter_name_cell.value, 'magenta'))
                    print(colored(defaulter_phone_no_cell_value, 'magenta'))

                    print(line * 136)

    print()

    return defaulter_name_lst, defaulter_phone_no_lst


def defaulter_msg(defaulter_name_lst, defaulter_phone_no_lst):
    print(colored("Which person Do you want to txt: ".title(), 'yellow').center(136, "-"))
    sent_to_name = input(" " * 67)
    while sent_to_name not in defaulter_name_lst:
        print(colored("No Buyer of that name exists. Re-enter".title(), 'red').center(136, "-"))
        sent_to_name = input(" " * 67)
    index = defaulter_name_lst.index(sent_to_name)
    sent_to_phone = defaulter_phone_no_lst[index]

    pywhatkit.sendwhatmsg_instantly(sent_to_phone, "You have not made Payments in time. Please pay")


def plaza_update():
    x = 1
    names_lst = []
    phone_lst = []
    for rows in range(5, sheet1.max_row + 1):
        print(colored(f"Person {x}", 'green'))
        for column in range(2, 8):
            name_cell = sheet1.cell(rows, 4)
            name_cell_value = name_cell.value
            phone_cell = sheet1.cell(rows, 6)
            phone_cell_value = phone_cell.value
            phone_cell_value = phone_cell_value.split("-")
            phone_cell_value = "".join(phone_cell_value)
        names_lst.append(name_cell_value)
        phone_lst.append(phone_cell_value)
        print(name_cell_value)
        print(phone_cell_value)
        x += 1

    print(colored("Which person Do you want to txt: ".title(), 'yellow').center(136, "-"))
    sent_to_name = input(" " * 67)

    while sent_to_name not in names_lst:
        print(colored("No Buyer of that name exists".title(), 'red').center(136, "-"))
        sent_to_name = input(" " * 67)
    sent_to_phone = phone_lst[names_lst.index(sent_to_name)]

    update_msg_content = input(
        colored(f"Write Update To be sent to {sent_to_name}.".title(), "yellow").center(136) + "\n".center(126))

    pywhatkit.sendwhatmsg_instantly(sent_to_phone, update_msg_content)


def all_units_display():
    x = 1
    display_lst = ["Type: ", "Floor: ", "Size: ", "Quarterly Installment: ", "Total Price: ", "Status: ", "Sold to "
                                                                                                          "username: "]
    for row in range(2, sheet2.max_row + 1):
        print(colored(f"Serial No {x}", 'green').center(136, "-"))
        y = 0
        for column in range(2, 9):
            display_cell = sheet2.cell(row, column)
            print(display_lst[y] + str(display_cell.value).title())
            y += 1
        x += 1
        print()


# Buyer Choice 1, Displays options.......... And returns if want to buy
def show_available_units():
    sold_product_list = []
    for row in range(2, sheet2.max_row + 1):
        sold_product = sheet2.cell(row, 7)

        sold_product_list.append(sold_product.value)

    print()
    for conversion in range(len(sold_product_list)):
        if sold_product_list[conversion] == "sold":
            sold_product_list[conversion] = 0
        else:
            sold_product_list[conversion] = 1
    i = 0
    about = ['Unit Type:', 'Unit Floor:', 'Unit Size:', 'Quarterly Installment:', 'Total Price:', 'Unit is']
    index = 0
    x = 1
    s_no_value_lst = []
    potential_sales_single_element = []

    for rows in range(2, sheet2.max_row + 1):
        if sold_product_list[index] == 1:
            print(colored(f"Option {x}", 'green'))
            x += 1

        for column in range(2, 8):

            if sold_product_list[index] == 1:
                s_no_cell = sheet2.cell(rows, 1)
                s_no_value = s_no_cell.value
                if s_no_value not in s_no_value_lst:
                    s_no_value_lst.append(s_no_value)
                available_options_cell = sheet2.cell(rows, column)
                available_options_value = available_options_cell.value
                print(f"{about[i]} {available_options_value}".title())
                potential_sales_single_element.append(available_options_value)
                i += 1

        print()
        index += 1

        i = 0

    buy_options_full_detail_lst = []
    a = 6
    for i in range(0, len(potential_sales_single_element), 6):
        buy_options_full_detail_lst.append(potential_sales_single_element[i:a - 1])
        a += 6

    print(colored("Would you want to buy any of these? (Option Number or 0 for no): ", 'yellow').center(136))
    buy_or_not = int(input(" " * 67))
    while buy_or_not < 0 or buy_or_not > len(buy_options_full_detail_lst):
        print(colored("Invalid Entry. Enter Option Number or 0 for no: ", 'red').center(136))
        buy_or_not = int(input(" " * 67))

    if buy_or_not != 0:

        return buy_options_full_detail_lst[buy_or_not - 1], s_no_value_lst[buy_or_not - 1]

    else:

        return None, None


def monthly_expense():
    check_error = True
    while check_error:
        print(colored("Removing Previous Data", 'cyan').center(136, "-"))
        print()
        txt_file = open("finances report.txt", "r+")
        txt_file.truncate(0)
        txt_file = open("finances report.txt", "a")
        try:
            print(colored("Input Expenses for the current month".title(), 'cyan').center(136, "-") + "\n" +
                  colored("labour_expense".title(), 'cyan').center(136, "-") + "\n" +
                  colored("raw material".title(), 'cyan').center(136, "-") + "\n" +
                  colored("services".title(), 'cyan').center(136, "-") + "\n" +
                  colored("combined salaries".title(), 'cyan').center(136, "-") + "\n" +
                  colored("[Separate using space]", 'cyan').center(136, "-"))
            labour_expense, raw_mateial, services, combined_salaries = input(" " * 60).split()
        except ValueError:
            print(colored("Since you did not follow Instructions. We are redirecting you back to the menu".title(),
                          'red').center(136, "-"))
            break

        mydate = datetime.now().date()

        txt_file.write("".center(136, "-") + "\n")
        txt_file.write(f"Bill for the Month of {mydate}".center(136) + "\n")
        txt_file.write(f"Labour Expense cost in PKR: ".ljust(36) + "".rjust(100) + f"{labour_expense}\n")
        txt_file.write(f"Raw Material cost in PKR: ".ljust(36) + "".rjust(100) + f"{raw_mateial}\n")
        txt_file.write(f"Services cost in PKR: ".ljust(36) + "".rjust(100) + f"{services}\n")
        txt_file.write(f"Combined Salaries in PKR: ".ljust(36) + "".rjust(100) + f"{services}\n")
        txt_file.write("".center(136, "-") + "\n")
        txt_file.close()

        print(colored("Report has been made and stored in file".title(), 'cyan').center(136, "-"))
        check_error = False


def read_file():
    txt_file = open("finances report.txt", "r+")

    lines = txt_file.readlines()
    if len(lines) != 0:
        for single_line in lines:
            print(single_line)
        txt_file.close()
    else:
        print(colored("Bill Has Not Been Made By Finance Officer.", 'red'))


# If wants to buy...... Updates xl file and updates xl date
def xl_update(s_no_value, username_given):
    installment_or_full = ""
    for rows in range(2, sheet2.max_row + 1):

        s_no_xl = sheet2.cell(rows, 1)

        s_no_xl_value = str(s_no_xl.value)
        xl_paid_percentage = sheet2.cell(int(s_no_xl_value) + 1, 10)

        if s_no_xl_value == s_no_value and int(xl_paid_percentage.value) != 100:

            xl_status = sheet2.cell(int(s_no_xl_value) + 1, 7)
            xl_status.value = "sold"

            xl_sold_to = sheet2.cell(int(s_no_value) + 1, 8)
            xl_sold_to.value = str(username_given)

            next_payment_to_be_made_date = sheet2.cell(int(s_no_xl_value) + 1, 12)
            next_payment_to_be_made_date.value = (datetime.now().date()) + xyz.timedelta(days=90)

            payment_made_date = sheet2.cell(int(s_no_xl_value) + 1, 11)
            payment_made_date.value = datetime.now().date()

            print(colored("How Will Yoy Pay for this?".title(), 'yellow').center(136) + "\n" +
                  colored("(1)Cash or (2)Cheque ".title(), 'yellow').center(136))
            cash_cheque = input()
            while not re.search('^[1-2]$', cash_cheque):
                cash_cheque = input(colored("Invalid Choice re-enter: ", 'red'))
            while installment_or_full.lower() != "full" and installment_or_full.lower() != "one":
                print(colored("Do you want to pay in full or a single installment? [Type: one or full] ".title(),
                              'yellow').center(136))

                installment_or_full = input(" " * 67)

            if installment_or_full.lower() == "full":

                xl_paid_amount = sheet2.cell(int(s_no_xl_value) + 1, 9)
                xl_total_price = sheet2.cell(int(s_no_xl_value) + 1, 6)

                xl_paid_percentage.value = "100"

                xl_paid_amount.value = xl_total_price.value

                xl_payment_status = sheet2.cell(int(s_no_xl_value) + 1, 13)
                xl_payment_status.value = "OK"

                next_payment_to_be_made_date = sheet2.cell(int(s_no_xl_value) + 1, 12)
                next_payment_to_be_made_date.value = "Paid in full"

                payment_made_date = sheet2.cell(int(s_no_xl_value) + 1, 11)
                payment_made_date.value = "Paid in full"

            elif installment_or_full.lower() == "one":

                xl_installment = sheet2.cell(int(s_no_xl_value) + 1, 5)

                xl_paid_amount = sheet2.cell(int(s_no_xl_value) + 1, 9)
                paid_amount = xl_paid_amount.value

                installment_value = xl_installment.value
                new_price = paid_amount + installment_value
                xl_paid_amount.value = new_price

                xl_paid_percentage = sheet2.cell(int(s_no_xl_value) + 1, 10)
                if xl_paid_percentage.value == "90":
                    next_payment_to_be_made_date = sheet2.cell(int(s_no_xl_value) + 1, 12)
                    next_payment_to_be_made_date.value = "Paid in full"

                    payment_made_date = sheet2.cell(int(s_no_xl_value) + 1, 11)
                    payment_made_date.value = "Paid in full"

                    xl_paid_percentage.value = "100"
                else:
                    new_percentage = int(xl_paid_percentage.value) + 10
                    xl_paid_percentage.value = new_percentage

                xl_payment_status = sheet2.cell(int(s_no_xl_value) + 1, 13)
                xl_payment_status.value = "OK"

            if cash_cheque == "1":
                print(line * 136)
                print(colored("You have paid using cash".title(), 'green').center(136))
                print(line * 136)

            if cash_cheque == "2":
                print(line * 136)
                print(colored("You have paid using cheque".title(), 'green').center(136))
                print(line * 136)
            if installment_or_full.lower() == "full":
                print(colored("You have Successfully Paid the Full amount.".title(), 'green').center(136))
                print(line * 136)
            elif installment_or_full.lower() == "one":
                print(colored("You have Successfully Paid a Single Installment.".title(), 'green').center(136))
                print(line * 136)
        elif int(xl_paid_percentage.value) == 100:
            print(colored("You have Already Paid Full Amount", 'red'))
    wb.save("transactions.xlsx")


def bought_units_display(sold_to_username_check, display_only):
    buy_or_not = 0
    show_stuff = False
    bought_product_lst = []
    for row in range(2, sheet2.max_row + 1):
        sold_to_username = sheet2.cell(row, 8)

        bought_product_lst.append(sold_to_username.value)

    print()
    for conversion in range(len(bought_product_lst)):
        # Ulta
        if bought_product_lst[conversion] == sold_to_username_check:
            bought_product_lst[conversion] = 1
        else:
            bought_product_lst[conversion] = 0
    i = 0
    about = ['Unit Type:', 'Unit Floor:', 'Unit Size:', 'Quarterly Installment:', 'Total Price:', 'Unit is']
    index = 0
    x = 1
    s_no_value_lst = []
    bought_item_single_element = []
    once = False
    for rows in range(2, sheet2.max_row + 1):

        if bought_product_lst[index] == 1:
            print(colored(f"Item {x}", 'magenta'))
            x += 1

        for column in range(2, 8):

            if bought_product_lst[index] == 1:
                s_no_cell = sheet2.cell(rows, 1)
                s_no_value = s_no_cell.value
                if s_no_value not in s_no_value_lst:
                    s_no_value_lst.append(s_no_value)
                available_options_cell = sheet2.cell(rows, column)
                available_options_value = available_options_cell.value
                print(f"{about[i]} {available_options_value}".title())
                bought_item_single_element.append(available_options_value)
                i += 1
                show_stuff = True

        if bought_product_lst[index] == 1:
            s_no_cell = sheet2.cell(rows, 1)
            s_no_value = s_no_cell.value
            if s_no_value not in s_no_value_lst:
                s_no_value_lst.append(s_no_value)
            payment_status_cell = sheet2.cell(rows, 13)
            payment_status_value = payment_status_cell.value
            print(colored(f"your payment status is: {payment_status_value}".title(), 'cyan'))

            i += 1
        if show_stuff and not once:
            print(line * 136)
            once = True
        index += 1

        i = 0

    buy_options_full_detail_lst = []
    a = 6

    for i in range(0, len(bought_item_single_element), 6):
        buy_options_full_detail_lst.append(bought_item_single_element[i:a - 1])
        a += 6
    if not show_stuff:
        print(colored("You have not bought anything", 'red').center(136, "-").title())

    if show_stuff and not display_only:
        print(colored("Do You want to make a payment? (Option Number or 0 for no): ", 'yellow').center(136))
        buy_or_not = int(input(" " * 67))
        while buy_or_not < 0 or buy_or_not > len(buy_options_full_detail_lst):
            print(colored("Invalid Entry. Enter Option Number or 0 for no: ", 'red').center(136))
            buy_or_not = int(input(" " * 67))

    if buy_or_not != 0:

        return buy_options_full_detail_lst[buy_or_not - 1], s_no_value_lst[buy_or_not - 1]

    else:

        return None, None


def sending_whatsapp_query(name):
    user_inquiry = input(colored("Please Ask Your Desired Q.".title(), 'magenta').center(136) + "\n" +
                         colored("The Q will Be sent to HR using whatsApp.", 'magenta').title().center(
                             136) + "\n".center(136))
    final_txt = user_inquiry + "\n" + f"This msg was sent by {name}"
    pywhatkit.sendwhatmsg_instantly("+923368055506", final_txt)
    txt_file = open("hr reply.txt", "a")
    txt_file.write(name + ",")
    txt_file.close()
